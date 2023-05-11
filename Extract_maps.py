import googlemaps
import openpyxl

# Define as informações da sua chave de API
gmaps = googlemaps.Client(key='Enter Your Google API')

# Obtém a profissão que deseja buscar
profissao = input('Digite a profissão que deseja buscar: ')

# Define as coordenadas e raio de busca
latitude = # Enter your Latitude
longitude = # Enter your Longitude
raio = # Enter the radius you want to search (in m)

# Faz a pesquisa no Google Maps com base nos parâmetros especificados
places = gmaps.places_nearby(location=(latitude, longitude), radius=raio, keyword=profissao)

# Cria um novo arquivo do Excel para salvar os resultados
workbook = openpyxl.Workbook()
sheet = workbook.active

# Adiciona os cabeçalhos das colunas
sheet['A1'] = 'Nome'
sheet['B1'] = 'Endereço'
sheet['C1'] = 'Telefone'

# Preenche o arquivo do Excel com os resultados da pesquisa
for index, place in enumerate(places['results']):
    # Obtém o ID do lugar
    place_id = place['place_id']

    # Faz uma chamada para obter mais detalhes sobre o lugar
    details = gmaps.place(place_id, fields=['name', 'formatted_address', 'formatted_phone_number'])

    # Obtém as informações necessárias
    nome = details['result']['name']
    endereco = details['result']['formatted_address']
    telefone = details['result'].get('formatted_phone_number', 'Não encontrado')
    sheet.cell(row=index+2, column=1).value = nome
    sheet.cell(row=index+2, column=2).value = endereco
    sheet.cell(row=index+2, column=3).value = telefone

# Salva o arquivo do Excel
workbook.save(filename='resultados.xlsx')
